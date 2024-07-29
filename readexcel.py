from openpyxl import load_workbook
from datetime import datetime

# Load the workbook and select the active worksheet
workbook = load_workbook('DATA.xlsx')
sheet = workbook.active

StudentDetails = []

# Iterate over rows and skip blank rows
for row in sheet.iter_rows(values_only=True):
    if row[0] == "S.NO.":
        continue

    if any(cell is not None for cell in row):  # Check if there is at least one non-None cell

        # Roll no and dob must not be none
        if row[3] is not None and row[4] is not None:

            date_obj = row[4]
            if isinstance(date_obj, datetime):
                formatted_date = date_obj.strftime('%m/%d/%Y')
            else:
                formatted_date = None

            # Print the row with formatted date
            StudentDetails.append({ "name": row[1], "rollno": row[3], "dob": formatted_date} )

print(StudentDetails)